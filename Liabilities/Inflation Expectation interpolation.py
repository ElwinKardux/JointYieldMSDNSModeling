# -*- coding: utf-8 -*-
"""
Created on Wed Jan 13 17:36:42 2021

@author: elwin
"""
from datetime import datetime, date, time, timedelta
from dateutil.relativedelta import relativedelta
from scipy.interpolate import CubicSpline
from scipy.interpolate import interp1d
from scipy.optimize import least_squares

import pandas as pd
import numpy as np
import xlrd
from scipy.interpolate import interp1d
from openpyxl import load_workbook
import matplotlib.pyplot as plt
plt.rcParams['axes.labelsize'] = 14
plt.rcParams['xtick.labelsize'] = 12
plt.rcParams['ytick.labelsize'] = 12
plt.rcParams['figure.figsize'] = 15, 10

def compound(x):
    #print (x)
    a = x.add(1).cumprod()
    a.Returns.iat[0] = 1
    return a


# wb_inflbond = load_workbook(filename="inflationbonddata.xlsx")
# #wb1 = xlrd.open_workbook('inflationbonddata.xlsx', 'Germany ILB')
# wb2 = load_workbook(filename="CF pattern EUR 2021 inflation assignment.xlsx")





data_cf = pd.read_excel(r'CF pattern EUR 2021 inflation assignment.xlsx')
data_infexp2021 = pd.read_excel(r'CF pattern EUR 2021 inflation assignment.xlsx', 'Sheet2')
data_ECB = pd.read_excel(r'ECB_data.xlsx')

cf = data_cf.iloc[2:,1]
ecb = data_ECB.iloc[3:,1]
ecb = ecb.reset_index(inplace = False)
ecb = ecb.iloc[:,1]

inf_r = data_infexp2021.iloc[0,:]

mat = np.array([1,2,3,4,5,6,7,8,9,10,15,20,30,40,50,100])
mat2 = np.arange(1,31)


# plt.xlabel('Time to Maturity (in years)')
# plt.ylabel('Best Inflation Estimates (%)')
# plt.plot(mat, inf_r, 'o', label = 'Best inflation estimate')
# plt.title('Inflation Points\nDate: 12/31/2020', fontsize = 15)
# plt.legend()
# plt.grid(linestyle = '--', linewidth = 1)


plt.xlabel('Time to Maturity (in years)')
plt.ylabel('ECB (%)')
plt.plot(mat2, ecb, 'o', label = 'ecb')
plt.title('ECB \nDate: 12/31/2020', fontsize = 15)
plt.legend()
plt.grid(linestyle = '--', linewidth = 1)


# =============================================================================
#  Cubic Spline?
# =============================================================================
cs = CubicSpline(mat, inf_r)
curve_points = np.linspace(min(mat),max(mat), 100)

# plt.subplot(1,2,1)
# plt.title('Inflation Curve from Cubic Spline\nDate: 12/31/2020', fontsize = 13)
# plt.xlabel('Time to Maturity (in months)')
# plt.ylabel('Inflation Rate (%)')
# plt.plot(mat, inf_r*100, 'o', label = 'Best inflation estimate')
# plt.plot(curve_points, cs(curve_points)*100, label="cubic spline")
# plt.grid(linestyle = '--', linewidth = 1)
# plt.legend()

cs_ecb = CubicSpline(mat2, ecb)
curve_points_ecb = np.linspace(min(mat2),65, 65)

plt.plot(1,2,1)
plt.title('ECB from Cubic Spline\nDate: 12/31/2020', fontsize = 13)
plt.xlabel('Time to Maturity (in months)')
plt.ylabel('ECB (%)')
plt.plot(mat2, ecb, 'o', label = 'ECB')
plt.plot(curve_points_ecb, cs_ecb(curve_points_ecb), label="cubic spline")
plt.grid(linestyle = '--', linewidth = 1)
#plt.legend()


infl_estim = cs(curve_points) #Best estimate inflation
infl_estim2 = np.ones(100) * 0.03
arr = np.arange(100)
index = 100

indexation = index * np.power(1+infl_estim,arr)
indexation2 = index *  np.power(1+infl_estim2,arr)
cf2 = cf * indexation[0:65]/100
cf3 = cf * indexation2[0:65]/100
cf_df = cf.to_frame()

Cashflows0 = np.sum(cf) #5.2 billion
Cashflows1 = np.sum(cf2) #6.96 billion
Cashflows2 = np.sum(cf3)
cf_df['indexation'] = 100


